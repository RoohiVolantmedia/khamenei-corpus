"""صفحه‌ی صادرات داده"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import streamlit as st
import sqlite3
import json
import csv
import io
from datetime import datetime
from utils.theme import render_header, render_footer

# مسیر دیتابیس — محلی یا سرور
def _resolve_db():
    from pathlib import Path as _P
    local = _P(__file__).resolve().parent.parent.parent / "database.db"
    if local.exists():
        return str(local)
    tmp = _P("/tmp/khamenei_db.db")
    if tmp.exists():
        return str(tmp)
    from utils.db import _get_db_path
    return str(_get_db_path())
DB_PATH = _resolve_db()

render_header("صادرات داده")
st.markdown("""
این صفحه برای **دانلود انبوه** اسناد پیکره طراحی شده است.
فیلتر بزنید، تعداد نتایج را ببینید، سپس در فرمت دلخواه دانلود کنید.

> برای جستجوی پیشرفته‌تر و دانلود تک‌سند به صفحه **جستجو** بروید.
""")
st.divider()

@st.cache_resource
def get_db():
    c = sqlite3.connect(DB_PATH, check_same_thread=False)
    c.execute("PRAGMA journal_mode=WAL")
    c.row_factory = sqlite3.Row
    return c
conn = get_db()

TIER_FA = {
    'tier1': 'tier1 — سخنرانی مستقیم',
    'tier2': 'tier2 — متن ویرایش‌شده',
    'tier3': 'tier3 — خلاصه / گزیده',
    'tier4': 'tier4 — گزارش راوی',
}

# ─── فیلترها ─────────────────────────────────────────────────────────────────
with st.expander("فیلترها", expanded=True):
    keyword = st.text_input("کلیدواژه (اختیاری)", key='ex_kw',
                            help="اگر خالی باشد، همه اسناد با فیلترهای دیگر صادر می‌شود")
    cc1, cc2, cc3 = st.columns(3)
    sel_tier = cc1.multiselect("نوع سند", ['tier1','tier2','tier3','tier4'],
                                format_func=lambda x: TIER_FA.get(x, x))
    sel_src  = cc1.multiselect("منبع", ['speech','speech_supplement','message','decree'])
    yr1 = cc2.number_input("از سال", 1356, 1404, 1356, step=1, key='ey1')
    yr2 = cc2.number_input("تا سال", 1356, 1404, 1404, step=1, key='ey2')
    max_n = cc3.selectbox("حداکثر نتایج", [100, 500, 1000, 5000, 42749],
                           format_func=lambda x: f"{x:,} سند")

# ─── کوئری ───────────────────────────────────────────────────────────────────
conds, params = [], []
if keyword.strip():
    conds.append("(d.title LIKE ? OR d.full_text LIKE ?)"); params += [f'%{keyword}%']*2
if sel_tier:
    conds.append(f"d.analytical_tier IN ({','.join('?'*len(sel_tier))})"); params.extend(sel_tier)
if sel_src:
    conds.append(f"d.content_source IN ({','.join('?'*len(sel_src))})"); params.extend(sel_src)
conds.append("(d.date_persian IS NULL OR d.date_persian BETWEEN ? AND ?)")
params.extend([f"{int(yr1)}/01/01", f"{int(yr2)}/12/29"])

where = "WHERE " + " AND ".join(conds)
join  = "LEFT JOIN doc_tags t ON d.doc_id=t.doc_id"

total = conn.execute(f"SELECT COUNT(*) FROM documents d {join} {where}", params).fetchone()[0]
export_n = min(total, max_n)
st.info(f"**{total:,} سند** با این فیلترها پیدا شد — صادرات **{export_n:,} سند**")

if total == 0:
    st.warning("نتیجه‌ای یافت نشد."); st.stop()

# ─── واکشی داده ──────────────────────────────────────────────────────────────
@st.cache_data(ttl=120)
def fetch_rows(_conn, where, params, max_n, join):
    rows_raw = _conn.execute(f"""
        SELECT d.doc_id, d.date_persian, d.analytical_tier, d.content_source,
               d.period_label, d.title, d.word_count, d.url,
               substr(d.full_text,1,300) AS snippet,
               d.full_text,
               t.tag_voice_type, t.tag_form_genre, t.tag_occasion,
               t.tag_topics, t.tag_regions, t.tag_audience, t.tag_tone
        FROM documents d {join} {where}
        ORDER BY d.date_persian DESC LIMIT ?
    """, params + (max_n,)).fetchall()
    return [dict(r) for r in rows_raw]

rows = fetch_rows(conn, where, tuple(params), max_n, join)

ts = datetime.now().strftime('%Y%m%d_%H%M%S')

# ─── CSV ──────────────────────────────────────────────────────────────────────
st.subheader("الف) CSV")
st.caption("مناسب برای Excel، R، Python و ابزارهای تحلیل داده")
buf = io.StringIO()
fields = ['doc_id','date_persian','analytical_tier','content_source','period_label',
          'title','word_count','url','tag_voice_type','tag_form_genre',
          'tag_occasion','tag_topics','tag_regions','tag_audience','tag_tone','snippet']
w = csv.DictWriter(buf, fieldnames=fields, extrasaction='ignore')
w.writeheader(); w.writerows(rows)
st.download_button("دانلود CSV", ('﻿' + buf.getvalue()).encode('utf-8'),
                   file_name=f"khamenei_{ts}.csv", mime="text/csv",
                   use_container_width=True)

# ─── Excel ────────────────────────────────────────────────────────────────────
st.subheader("ب) Excel")
st.caption("شامل ۴ Sheet: اطلاعات / متادیتا / متن کامل / تگ ۷ بعدی")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()

    ws1 = wb.active; ws1.title = 'اطلاعات'; ws1.sheet_view.rightToLeft = True
    for r in [['کلیدواژه', keyword or '—'],
              ['نوع سند', str([TIER_FA.get(t,t) for t in sel_tier]) if sel_tier else 'همه'],
              ['بازه سال', f"{int(yr1)}–{int(yr2)}"],
              ['تعداد نتایج', len(rows)],
              ['تاریخ صادرات', datetime.now().strftime('%Y-%m-%d %H:%M')]]:
        ws1.append(r)

    ws2 = wb.create_sheet('متادیتا'); ws2.sheet_view.rightToLeft = True
    meta_cols = ['doc_id','date_persian','analytical_tier','content_source',
                 'period_label','title','word_count','url',
                 'tag_voice_type','tag_form_genre','tag_occasion',
                 'tag_topics','tag_regions','tag_audience','tag_tone']
    ws2.append(meta_cols)
    for row in rows: ws2.append([row.get(c,'') for c in meta_cols])

    ws3 = wb.create_sheet('متن کامل'); ws3.sheet_view.rightToLeft = True
    ws3.append(['doc_id','date_persian','title','full_text'])
    for row in rows:
        ws3.append([row.get('doc_id',''), row.get('date_persian',''),
                    row.get('title',''), row.get('full_text','')])

    ws4 = wb.create_sheet('تگ ۷ بعدی'); ws4.sheet_view.rightToLeft = True
    tag_cols = ['doc_id','tag_voice_type','tag_form_genre','tag_audience',
                'tag_occasion','tag_topics','tag_regions','tag_tone']
    ws4.append(tag_cols)
    for row in rows: ws4.append([row.get(c,'') for c in tag_cols])

    hf = Font(bold=True, name='Vazirmatn')
    hfill = PatternFill('solid', fgColor='E8E8E8')
    for ws in [ws1, ws2, ws3, ws4]:
        for cell in ws[1]:
            cell.font = hf; cell.fill = hfill
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name='Vazirmatn', size=9)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    xl_buf = io.BytesIO(); wb.save(xl_buf)
    st.download_button("دانلود Excel", xl_buf.getvalue(),
                       file_name=f"khamenei_{ts}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=True)
except Exception as e:
    st.error(f"خطای Excel: {e}")

# ─── JSON ─────────────────────────────────────────────────────────────────────
st.subheader("پ) JSON")
st.caption("مناسب برای برنامه‌نویسی و API")
json_rows = [{k:v for k,v in r.items() if k != 'full_text'} for r in rows]
st.download_button("دانلود JSON (بدون متن کامل)",
                   json.dumps(json_rows, ensure_ascii=False, indent=2).encode('utf-8'),
                   file_name=f"khamenei_{ts}.json", mime="application/json",
                   use_container_width=True)
json_full = rows
st.download_button("دانلود JSON (با متن کامل)",
                   json.dumps(json_full, ensure_ascii=False, indent=2).encode('utf-8'),
                   file_name=f"khamenei_full_{ts}.json", mime="application/json",
                   use_container_width=True)

# ─── پیش‌نمایش ────────────────────────────────────────────────────────────────
st.divider()
st.subheader("پیش‌نمایش (۵۰ سند اول)")
import pandas as pd
df = pd.DataFrame([{c: r.get(c,'') for c in
    ['doc_id','date_persian','analytical_tier','title','word_count','tag_form_genre','tag_topics']}
    for r in rows[:50]])
st.dataframe(df, use_container_width=True, height=320)

render_footer()
